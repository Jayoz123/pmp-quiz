#!/usr/bin/env python3
"""
Konwertuje przetłumaczony plik questions_pl.xlsx do questions.json
Użycie: python tools/convert_to_json.py [input.xlsx] [output.json]
"""
import openpyxl
import json
import sys
import os
import random


LETTER_TO_INDEX = {'A': 0, 'B': 1, 'C': 2, 'D': 3}


def convert(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
    required = ['ID', 'Domain', 'Question', 'Answer_A', 'Answer_B', 'Answer_C', 'Answer_D', 'Correct', 'Explanation']
    for req in required:
        if req not in headers:
            print(f'ERROR: Missing column "{req}". Found columns: {headers}')
            sys.exit(1)

    col = {h: i for i, h in enumerate(headers)}
    questions = []
    errors = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[col['Question']]:
            continue

        correct_letter = str(row[col['Correct']]).strip().upper() if row[col['Correct']] else ''
        if correct_letter not in LETTER_TO_INDEX:
            print(f'Row {row_num}: Invalid Correct value "{correct_letter}", skipping')
            errors += 1
            continue

        answers_ordered = [
            str(row[col['Answer_A']] or '').strip(),
            str(row[col['Answer_B']] or '').strip(),
            str(row[col['Answer_C']] or '').strip(),
            str(row[col['Answer_D']] or '').strip(),
        ]
        correct_text = answers_ordered[LETTER_TO_INDEX[correct_letter]]

        # Shuffle answers so correct position is random in the JSON
        indexed = [(t, t == correct_text) for t in answers_ordered]
        random.shuffle(indexed)
        shuffled_answers = [t for t, _ in indexed]
        correct_index = next(i for i, (_, is_c) in enumerate(indexed) if is_c)

        questions.append({
            'id': int(row[col['ID']]) if row[col['ID']] else row_num - 1,
            'domain': str(row[col['Domain']] or 'General').strip(),
            'question': str(row[col['Question']]).strip(),
            'answers': shuffled_answers,
            'correct': correct_index,
            'explanation': str(row[col['Explanation']] or '').strip(),
        })

    if errors:
        print(f'WARNING: {errors} rows skipped due to errors')

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(questions, f, ensure_ascii=False, indent=2)

    print(f'Converted {len(questions)} questions → {output_path}')
    return questions


def main():
    input_path  = sys.argv[1] if len(sys.argv) > 1 else 'questions_pl.xlsx'
    output_path = sys.argv[2] if len(sys.argv) > 2 else 'pmp-quiz-app/questions.json'
    if not os.path.exists(input_path):
        print(f'ERROR: {input_path} not found')
        sys.exit(1)
    convert(input_path, output_path)


if __name__ == '__main__':
    main()
