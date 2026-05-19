#!/usr/bin/env python3
"""
Ekstrahuje pytania z plikow PDF PMP i zapisuje do questions_en.xlsx.
Uzycie: python tools/extract_questions.py

Obsluguje dwa formaty PDF:
- Format 1 (PDF1): numerowane pytania z a./b./c./d. i markerem [Ans]
- Format 2 (PDF2): sekcje Q i A z "Question N" + A)/B)/C)/D) + "(Correct Answer)"
"""
import pdfplumber
import openpyxl
import re
import random
import os
import sys

PDF_FILES = [
    "404451735-Questions-and-Answers-PMP-Exam-Prep.pdf",
    "983625620-PMP-Exam-Prep-2025-2026-All-In-One-Guide-to-Passing-With-Confidence-Including-1-100-Practice-Test-Proven-Strategies-to-Get-Publication-NewGrade.pdf",
]

# Page ranges for each practice test in PDF2 (0-indexed start, exclusive end)
# (q_start, q_end, ans_start, ans_end)
PDF2_SECTIONS = [
    (172, 252, 251, 306),
    (304, 390, 389, 446),
    (444, 510, 509, 552),
    (550, 595, 594, 626),
    (624, 665, 664, 695),
    (693, 745, 744, 775),
]

PMP_DOMAINS = ['Risk', 'Cost', 'Schedule', 'Scope', 'Quality', 'Resource',
               'Communications', 'Stakeholder', 'Procurement', 'Integration',
               'People', 'Process', 'Business Environment']

DOMAIN_KEYWORDS = {
    'Risk': ['risk', 'threat', 'opportunity', 'probability', 'impact', 'mitigation'],
    'Cost': ['cost', 'budget', 'EVM', 'earned value', 'CPI', 'CV', 'BAC', 'EAC'],
    'Schedule': ['schedule', 'SPI', 'SV', 'critical path', 'float', 'slack', 'CPM'],
    'Scope': ['scope', 'WBS', 'requirements', 'deliverable', 'change request'],
    'Quality': ['quality', 'QA', 'QC', 'defect', 'audit', 'process improvement'],
    'Resource': ['resource', 'team', 'RACI', 'staffing', 'training', 'HR'],
    'Communications': ['communication', 'report', 'stakeholder', 'message', 'channel'],
    'Stakeholder': ['stakeholder', 'engagement', 'register', 'influence', 'interest'],
    'Procurement': ['procurement', 'contract', 'vendor', 'RFP', 'SOW', 'make-or-buy'],
    'Integration': ['integration', 'charter', 'project plan', 'change control', 'lessons'],
    'People': ['servant leader', 'emotional intelligence', 'conflict', 'motivation', 'team'],
    'Process': ['process group', 'knowledge area', 'initiating', 'planning', 'executing'],
    'Business Environment': ['compliance', 'governance', 'benefit', 'strategy', 'organization'],
}


def infer_domain(question_text, explanation_text=''):
    text = (question_text + ' ' + explanation_text).lower()
    scores = {domain: 0 for domain in PMP_DOMAINS}
    for domain, keywords in DOMAIN_KEYWORDS.items():
        for kw in keywords:
            if kw.lower() in text:
                scores[domain] += 1
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else 'General'


def extract_pages(path, start, end):
    """Extract text from a page range (0-indexed, exclusive end)."""
    pages = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages[start:end]:
            text = page.extract_text()
            if text:
                pages.append(text)
    return '\n'.join(pages)


def shuffle_answers(choices, correct_letter):
    """Randomly place correct answer in A/B/C/D slot, shuffle others."""
    other_letters = [l for l in ['A', 'B', 'C', 'D'] if l in choices and l != correct_letter]
    random.shuffle(other_letters)
    all_cols = ['A', 'B', 'C', 'D']
    correct_col = random.choice(all_cols)
    remaining_cols = [c for c in all_cols if c != correct_col]
    mapping = {correct_col: correct_letter}
    for col, src in zip(remaining_cols, other_letters):
        mapping[col] = src
    answers = {col: choices.get(mapping.get(col, ''), '') for col in all_cols}
    return answers, correct_col


def parse_questions_format1(text):
    """
    Format 1: numbered questions (1. 2. ...) with lowercase a./b./c./d. choices
    and [Ans] marker after the correct answer choice.
    """
    questions = []

    block_pattern = re.compile(
        r'(?:^|\n)(\d+)[.)]\s*\n?(.+?)(?=(?:^|\n)\d+[.)]\s|\Z)',
        re.DOTALL | re.MULTILINE
    )
    lc_choice_pattern = re.compile(
        r'(?:^|\n)([a-d])\.\s+(.+?)(?=\n[a-d]\.\s|\Z)',
        re.DOTALL | re.MULTILINE
    )

    for match in block_pattern.finditer(text):
        block = match.group(2).strip()
        lines = block.split('\n')

        question_lines = []
        for line in lines:
            if re.match(r'[a-d]\.\s', line):
                break
            question_lines.append(line.strip())

        question_text = ' '.join(q for q in question_lines if q).strip()
        if len(question_text) < 20:
            continue

        if '[Ans]' not in block and '[ans]' not in block:
            continue

        choices_raw = lc_choice_pattern.findall(block)
        if len(choices_raw) < 2:
            continue

        choices = {}
        correct_letter = None
        for letter, choice_text in choices_raw:
            clean = re.sub(r'\s+', ' ', choice_text).strip()
            if re.search(r'\[Ans\]', clean, re.IGNORECASE):
                clean = re.sub(r'\s*\[Ans\]', '', clean, flags=re.IGNORECASE).strip()
                correct_letter = letter.upper()
            choices[letter.upper()] = clean

        if correct_letter is None or correct_letter not in choices:
            continue

        answers, correct_col = shuffle_answers(choices, correct_letter)

        questions.append({
            'question': question_text,
            'answer_a': answers['A'],
            'answer_b': answers['B'],
            'answer_c': answers['C'],
            'answer_d': answers['D'],
            'correct': correct_col,
            'explanation': '',
        })

    return questions


def parse_question_section(text):
    """
    Parse question section pages.
    Returns dict: qnum -> {'text': ..., 'choices': {A:..., B:..., ...}}
    """
    q_block_pattern = re.compile(
        r'Question\s+(\d+)\s*\n(.+?)(?=Question\s+\d+|\Z)',
        re.DOTALL | re.IGNORECASE
    )
    choice_pattern = re.compile(
        r'(?:^|\n)\s*([A-D])\)\s+(.+?)(?=(?:^|\n)\s*[A-D]\)|Question\s+\d+|\Z)',
        re.DOTALL | re.MULTILINE
    )

    questions = {}
    for m in q_block_pattern.finditer(text):
        qnum = int(m.group(1))
        block = m.group(2).strip()

        if re.search(r'\(Correct\s+Answer', block, re.IGNORECASE):
            continue

        choices_found = choice_pattern.findall(block)
        if len(choices_found) < 2:
            continue

        lines = block.split('\n')
        q_lines = []
        for line in lines:
            if re.match(r'\s*[A-D]\)', line):
                break
            q_lines.append(line.strip())
        q_text = ' '.join(q for q in q_lines if q).strip()

        choices = {}
        for letter, choice_text in choices_found:
            clean = re.sub(r'\s+', ' ', choice_text).strip()
            clean = re.sub(r'\s*\(Correct\s+Answer.*', '', clean, flags=re.IGNORECASE).strip()
            choices[letter.upper()] = clean

        if len(q_text) >= 15 and len(choices) >= 2:
            questions[qnum] = {'text': q_text, 'choices': choices}

    return questions


def parse_answer_section_numbered(text):
    """
    Parse answer section with 'N.' format and X) answer (Correct Answer).
    Returns dict: qnum -> {'correct': letter, 'explanation': text}
    """
    ans_block_pattern = re.compile(
        r'(?:^|\n)(\d+)[.]\s*\n?\s*([A-D])\)\s+(.+?)\s*\(Correct\s+Answers?\)',
        re.DOTALL | re.MULTILINE | re.IGNORECASE
    )
    explanation_pattern = re.compile(
        r'Explanation[:\s]+(.+?)(?=(?:^|\n)\d+[.]\s|\Z)',
        re.DOTALL | re.IGNORECASE
    )

    answers = {}
    for m in ans_block_pattern.finditer(text):
        qnum = int(m.group(1))
        correct_letter = m.group(2).upper()
        rest = text[m.end():]
        exp_match = explanation_pattern.match(rest.lstrip())
        if not exp_match:
            exp_match = explanation_pattern.search(rest[:600])
        explanation = ''
        if exp_match:
            explanation = re.sub(r'\s+', ' ', exp_match.group(1)).strip()[:400]
        answers[qnum] = {'correct': correct_letter, 'explanation': explanation}

    return answers


def parse_answer_section_keyword(text):
    """
    Parse answer section with 'Question N' headers and X) answer (Correct Answer).
    Returns dict: qnum -> {'correct': letter, 'explanation': text}
    """
    q_block_pattern = re.compile(
        r'Question\s+(\d+)\s*\n(.+?)(?=Question\s+\d+|\Z)',
        re.DOTALL | re.IGNORECASE
    )
    correct_pattern = re.compile(
        r'([A-D])\)\s+(.+?)\s*\(Correct\s+Answers?\)',
        re.IGNORECASE | re.DOTALL
    )
    explanation_pattern = re.compile(
        r'Explanation[:\s]+(.+?)(?=Question\s+\d+|\Z)',
        re.DOTALL | re.IGNORECASE
    )

    answers = {}
    for m in q_block_pattern.finditer(text):
        qnum = int(m.group(1))
        block = m.group(2).strip()

        if not re.search(r'\(Correct\s+Answer', block, re.IGNORECASE):
            continue

        correct_matches = list(correct_pattern.finditer(block))
        if not correct_matches:
            continue

        correct_letter = correct_matches[0].group(1).upper()
        exp_match = explanation_pattern.search(block)
        explanation = ''
        if exp_match:
            explanation = re.sub(r'\s+', ' ', exp_match.group(1)).strip()[:400]

        answers[qnum] = {'correct': correct_letter, 'explanation': explanation}

    return answers


def join_questions_and_answers(q_dict, a_dict):
    """Join question dict and answer dict by question number."""
    questions = []
    for qnum in sorted(q_dict.keys()):
        if qnum not in a_dict:
            continue
        q = q_dict[qnum]
        a = a_dict[qnum]
        correct_letter = a['correct']
        choices = q['choices']

        if correct_letter not in choices:
            if len(choices) >= 2:
                correct_letter = list(choices.keys())[0]
            else:
                continue

        answers, correct_col = shuffle_answers(choices, correct_letter)

        questions.append({
            'question': q['text'],
            'answer_a': answers['A'],
            'answer_b': answers['B'],
            'answer_c': answers['C'],
            'answer_d': answers['D'],
            'correct': correct_col,
            'explanation': a.get('explanation', ''),
        })

    return questions


def extract_pdf2_section(pdf_path, q_start, q_end, ans_start, ans_end):
    """Extract one practice test section from PDF2."""
    q_text = extract_pages(pdf_path, q_start, q_end)
    # Include the last Q page in ans range (answers often start on last Q page)
    ans_raw = extract_pages(pdf_path, ans_start, ans_end)

    q_dict = parse_question_section(q_text)

    # Trim ans_text to start from "Correct Answers" header if present
    ca_idx = ans_raw.find('Correct Answers')
    if ca_idx == -1:
        ca_idx = ans_raw.find('CORRECT ANSWERS')
    ans_text = ans_raw[ca_idx:] if ca_idx >= 0 else ans_raw

    # Try numbered format first (N. X) answer), fall back to keyword (Question N)
    a_dict = parse_answer_section_numbered(ans_text)
    if len(a_dict) < 10:
        a_dict = parse_answer_section_keyword(ans_text)

    # Supplement with keyword format if numbered gave partial results
    if len(a_dict) < len(q_dict) // 2:
        a_dict2 = parse_answer_section_keyword(ans_text)
        for k, v in a_dict2.items():
            if k not in a_dict:
                a_dict[k] = v

    questions = join_questions_and_answers(q_dict, a_dict)
    return questions, len(q_dict), len(a_dict)


def remove_duplicates(questions):
    seen = set()
    unique = []
    for q in questions:
        key = re.sub(r'\s+', ' ', q['question'][:80].lower().strip())
        if key not in seen:
            seen.add(key)
            unique.append(q)
    return unique


def save_to_excel(questions, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Questions'
    headers = ['ID', 'Domain', 'Question', 'Answer_A', 'Answer_B', 'Answer_C', 'Answer_D', 'Correct', 'Explanation']
    ws.append(headers)

    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 60
    for col in ['D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 35
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 60

    for i, q in enumerate(questions, start=1):
        domain = infer_domain(q['question'], q['explanation'])
        ws.append([
            i, domain,
            q['question'], q['answer_a'], q['answer_b'], q['answer_c'], q['answer_d'],
            q['correct'], q['explanation'],
        ])
        for col in [3, 4, 5, 6, 7, 9]:
            ws.cell(row=i + 1, column=col).alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(output_path)
    print(f'Saved {len(questions)} questions to {output_path}')


def main():
    all_questions = []

    # PDF 1: Format with [Ans] marker
    pdf1 = PDF_FILES[0]
    if os.path.exists(pdf1):
        print(f'Processing {pdf1}...')
        text = extract_pages(pdf1, 0, 9999)
        questions = parse_questions_format1(text)
        print(f'  Extracted {len(questions)} questions ([Ans] format)')
        all_questions.extend(questions)
    else:
        print(f'WARNING: {pdf1} not found, skipping')

    # PDF 2: Format with Question N / (Correct Answer)
    pdf2 = PDF_FILES[1]
    if os.path.exists(pdf2):
        print(f'Processing {pdf2}...')
        total_from_pdf2 = 0
        for idx, (q_start, q_end, ans_start, ans_end) in enumerate(PDF2_SECTIONS, 1):
            print(f'  Section {idx} (Q pages {q_start+1}-{q_end}, A pages {ans_start+1}-{ans_end})...', end='', flush=True)
            try:
                section_qs, n_q, n_a = extract_pdf2_section(pdf2, q_start, q_end, ans_start, ans_end)
                print(f' {n_q} Qs + {n_a} As -> {len(section_qs)} matched')
                all_questions.extend(section_qs)
                total_from_pdf2 += len(section_qs)
            except Exception as e:
                print(f' ERROR: {e}')
        print(f'  Total from PDF2: {total_from_pdf2} questions')
    else:
        print(f'WARNING: {pdf2} not found, skipping')

    all_questions = remove_duplicates(all_questions)
    print(f'After deduplication: {len(all_questions)} questions')

    if not all_questions:
        print('ERROR: No questions extracted. Check PDF format.')
        sys.exit(1)

    save_to_excel(all_questions, 'questions_en.xlsx')
    print('\nNext step: run tools/convert_to_json.py to generate questions.json')


if __name__ == '__main__':
    main()
